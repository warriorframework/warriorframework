'''
Copyright 2017, Fujitsu Network Communications, Inc.
Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at
http://www.apache.org/licenses/LICENSE-2.0
Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

"""
This module has class and methods required to parse from and 
write to an xl workbook
"""

from collections import OrderedDict

class Wxl(object):
    """Wxl class has methods required to parse/write
    from/to xl sheet"""

    def __init__(self):
        """constructor for WRest """
        self.req = None
        self.import_openpyxl()

    def import_openpyxl(self):
        """Import the requests module """
        try:
            import openpyxl
            from openpyxl import load_workbook
        except ImportError:
            print("openpyxl module is not installed"\
                       "Please install openpyxl module to"\
                       "perform any activities related to parsing xl sheets")
        else:
            self.openpyxl = openpyxl
            self.load = load_workbook

            
    def load_workbook(self, wb_location):
        """
        Load an existing xl workbook 
        """

        value = None
        try:
            wb = self.load(wb_location)
        except Exception as err:
            print err
            print("Error loading workbook, check if file exists")
        else:
            value = wb
        return value

    def get_row_value_dict(self, worksheet, row_range):
        """
        """
        row_val_dict = OrderedDict()
        for row in worksheet.iter_rows(row_range):
            for cell in row:
                cell_id = str(cell.column) + str(cell.row)
                row_val_dict[cell_id] = cell.value
        return row_val_dict
    
    

    def get_row_for_value(self, worksheet, value, col, row_range):
        """
        Takes a value, column and row_range as input and returns the 
        row_num that has the matching value, returns None if no match
        found
        """
        found = False
        for row in worksheet.iter_rows(row_range):
            if found: break
            for cell in row:
                if found: break
                cell_value = cell.value
                if str(cell_value) == value:
                    return_value = str(cell.row)                    
                    found = True 
                else:
                    return_value = None
                    
        
        return return_value





                
